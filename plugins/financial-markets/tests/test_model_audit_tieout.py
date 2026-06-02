"""Model-audit artifact and issue-log regressions."""

from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "skills" / "model-audit-tieout" / "scripts" / "audit_workbook.py"


class ModelAuditTieoutTests(unittest.TestCase):
    def test_substantive_audit_uses_standalone_html_and_audit_verdict_language(self) -> None:
        skill = (ROOT / "skills" / "model-audit-tieout" / "SKILL.md").read_text(encoding="utf-8")
        dashboard_map = (
            ROOT / "skills" / "model-audit-tieout" / "references" / "dashboard-map.md"
        ).read_text(encoding="utf-8")
        templates = (
            ROOT / "skills" / "model-audit-tieout" / "references" / "output-templates.md"
        ).read_text(encoding="utf-8")
        taxonomy = (
            ROOT / "skills" / "model-audit-tieout" / "references" / "issue-taxonomy.md"
        ).read_text(encoding="utf-8")
        policy = (ROOT / "shared" / "deliverable-intake-policy.md").read_text(encoding="utf-8")

        for phrase in [
            "../../shared/html-artifact-standard.md",
            "polished standalone HTML model-audit report",
            "only when the user explicitly requests a standardized dashboard",
            "Do not use for portfolio action until remediated and re-audited",
            "Illustrative audit sensitivity",
            "linked valuation/scenario decision output",
            "local headless-browser screenshots",
            "mention supporting audit files without linking them unless the user asks",
            "external price anchor is unavailable",
        ]:
            self.assertIn(phrase, skill)

        for finding_type in [
            "Formula/control defect",
            "Source contradiction",
            "Unsupported assumption",
            "Missing forecast refresh",
            "Missing decision output",
            "Not comparable without bridge",
        ]:
            self.assertIn(finding_type, skill)
            self.assertIn(finding_type, taxonomy)

        self.assertIn(
            "finding count alone does not require a standardized dashboard", dashboard_map
        )
        self.assertIn("Illustrative audit sensitivity", templates)
        self.assertIn("polished standalone HTML model-audit report", policy)

    def test_audit_helper_outputs_html_issue_workbook_and_hidden_support(self) -> None:
        try:
            from openpyxl import Workbook
        except Exception as exc:  # pragma: no cover
            self.skipTest(f"openpyxl unavailable: {exc}")

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook = tmp_path / "audit_me.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Inputs"
            ws["A1"] = "Source date"
            ws["B2"] = "=TODAY()+1.23"
            hidden = wb.create_sheet("HiddenData")
            hidden.sheet_state = "hidden"
            hidden["A1"] = 42
            wb.save(workbook)

            out_dir = tmp_path / "audit_output"
            result = subprocess.run(
                [sys.executable, str(SCRIPT), str(workbook), "--out-dir", str(out_dir)],
                cwd=str(SCRIPT.parent),
                text=True,
                capture_output=True,
                check=False,
            )
            self.assertEqual(0, result.returncode, result.stderr + result.stdout)

            self.assertTrue((out_dir / "model_audit_report.html").exists())
            self.assertTrue((out_dir / "model_audit_issues.xlsx").exists())
            self.assertTrue((out_dir / "support" / "model_audit_findings.csv").exists())
            self.assertTrue((out_dir / "support" / "formula_exception_log.csv").exists())
            self.assertTrue((out_dir / "support" / "source_tieout_ledger.csv").exists())
            findings = json.loads(
                (out_dir / "logs" / "model_audit_findings.json").read_text(encoding="utf-8")
            )
            findings_text = json.dumps(findings)
            for phrase in [
                "first visible sheet is not a Cover tab",
                "sheet is hidden",
                "volatile function",
                "numeric literal",
            ]:
                self.assertIn(phrase, findings_text)
            manifest = json.loads((out_dir / "manifest.json").read_text(encoding="utf-8"))
            self.assertTrue(
                str(manifest["primary_human_deliverable"]).endswith("model_audit_report.html")
            )
            self.assertTrue(
                any(
                    item["path"].endswith("model_audit_issues.xlsx")
                    for item in manifest["companion_deliverables"]
                )
            )
            self.assertFalse(manifest["support_artifacts_user_visible_default"])
            self.assertTrue(manifest["support_artifacts"])
            self.assertTrue(
                all(not item["user_visible_default"] for item in manifest["support_artifacts"])
            )

    def test_update_cover_and_suffixed_cover_are_accepted_as_first_visible_cover_sheet(
        self,
    ) -> None:
        try:
            from openpyxl import Workbook
        except Exception as exc:  # pragma: no cover
            self.skipTest(f"openpyxl unavailable: {exc}")

        for title in ("Update_Cover", "Update_Cover_2"):
            with self.subTest(title=title), tempfile.TemporaryDirectory() as tmp:
                tmp_path = Path(tmp)
                workbook = tmp_path / "control_pack.xlsx"
                wb = Workbook()
                ws = wb.active
                ws.title = title
                ws["A1"] = "Bottom Line"
                ws["A2"] = "Estimate change and valuation change are blocked pending rebuild."
                ws["A3"] = "Revenue period and latest value chart source."
                ws["A4"] = "Source posture, stale items, and workbook map."
                wb.save(workbook)

                out_dir = tmp_path / "audit_output"
                result = subprocess.run(
                    [sys.executable, str(SCRIPT), str(workbook), "--out-dir", str(out_dir)],
                    cwd=str(SCRIPT.parent),
                    text=True,
                    capture_output=True,
                    check=False,
                )

                self.assertEqual(0, result.returncode, result.stderr + result.stdout)
                findings = json.loads(
                    (out_dir / "logs" / "model_audit_findings.json").read_text(encoding="utf-8")
                )
                findings_text = json.dumps(findings)
                self.assertNotIn("first visible sheet is not a Cover tab", findings_text)
                self.assertNotIn("cover/dashboard may be missing", findings_text)


if __name__ == "__main__":
    unittest.main()
